import urllib.request
import urllib.parse
import csv
import io
import re
import time
import openpyxl
from datetime import datetime, date, timedelta
from database import SessionLocal
import models
from gps_service import OFFICIAL_DRIVERS_BY_CLEAN_PLATE

DEFAULT_SHEET_AUGUST_ID = "1p0B1bx_yUM6BfW2D88P-Jgra3sBSfqHEM_Op35WxSpI"
DEFAULT_SHEET_SEPTEMBER_ID = "1NSNhOIO--PEx9W3u5O2I-w1bvl4kN3TTh9qXfiXb2U0"

AUGUST_DAY_GIDS = {
    "01": "2036778103", "02": "1298495420", "03": "349372223",
    "04": "1940989344", "05": "685790695",  "06": "775628549",
    "07": "426462719",  "08": "1517596825", "09": "1801267674",
    "10": "1002307521", "11": "119854746",  "12": "2029705886",
    "13": "1994784733", "14": "1815525547", "15": "917527663",
    "16": "1830588414", "17": "1564883446", "18": "864147041",
    "19": "1992019904", "20": "774136979",  "21": "1406859599",
    "22": "1408821946", "23": "673756214",  "24": "1369715783",
    "25": "1738747493", "26": "632069695",  "27": "1126602330",
    "28": "893952771",  "29": "1753303657",  "30": "1041617404",
    "31": "1753303657"
}

SHEET2_VEHICLES_URL = "https://docs.google.com/spreadsheets/d/1SzlPdtSjhBeqvFlZq5WiznFOVtF6SEvQ-5xKFF6I8Es/export?format=csv"
SHEET2_MAINTENANCE_URL = "https://docs.google.com/spreadsheets/d/1SzlPdtSjhBeqvFlZq5WiznFOVtF6SEvQ-5xKFF6I8Es/export?format=csv&gid=1041617404"

class GoogleSheetsSyncService:
    def __init__(self):
        self.cached_trips = {}          # date_str -> (timestamp, result_dict)
        self.cached_xlsx = {}           # sheet_id -> (timestamp, bytes)
        self.cached_maintenance = None  # (timestamp, list)
        self.cached_vehicles = None     # (timestamp, dict)
        self.last_sync_time = None

    def extract_sheet_id(self, url_or_id: str) -> str:
        if not url_or_id:
            return ""
        url_or_id = url_or_id.strip()
        m = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url_or_id)
        if m:
            return m.group(1)
        return url_or_id

    def get_sheet_id_for_month(self, month_str: str) -> str:
        db = SessionLocal()
        try:
            setting = db.query(models.SystemSetting).filter(
                models.SystemSetting.key == f"sheet_url_month_{month_str}"
            ).first()
            if setting and setting.value:
                return self.extract_sheet_id(setting.value)
        except Exception:
            pass
        finally:
            db.close()

        if month_str == "09":
            return DEFAULT_SHEET_SEPTEMBER_ID
        return DEFAULT_SHEET_AUGUST_ID

    def set_sheet_url_for_month(self, month_str: str, sheet_url: str):
        sheet_id = self.extract_sheet_id(sheet_url)
        db = SessionLocal()
        try:
            setting = db.query(models.SystemSetting).filter(
                models.SystemSetting.key == f"sheet_url_month_{month_str}"
            ).first()
            if not setting:
                setting = models.SystemSetting(
                    key=f"sheet_url_month_{month_str}",
                    value=sheet_url.strip(),
                    description=f"Google Sheet URL xe nhận hàng tháng {month_str}"
                )
                db.add(setting)
            else:
                setting.value = sheet_url.strip()
            db.commit()
            # Invalidate caches
            self.cached_trips.clear()
            self.cached_xlsx.clear()
        finally:
            db.close()

    def _get_xlsx_bytes(self, sheet_id: str) -> bytes:
        now = time.time()
        if sheet_id in self.cached_xlsx:
            ts, data = self.cached_xlsx[sheet_id]
            if now - ts < 600: # 10 phút cache XLSX
                return data

        try:
            xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            req = urllib.request.Request(xlsx_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = resp.read()
            self.cached_xlsx[sheet_id] = (now, data)
            return data
        except Exception as e:
            print(f"Error fetching XLSX for sheet {sheet_id}: {e}")
            raise e

    def fetch_daily_trips(self, target_date=None):
        today_obj = date.today()
        today_str = today_obj.strftime("%Y-%m-%d")
        t_date = target_date or today_str

        # Kiểm tra in-memory cache để chuyển ngày siêu tốc (< 5ms)
        now = time.time()
        if t_date in self.cached_trips:
            ts, cached_res = self.cached_trips[t_date]
            if now - ts < 300: # 5 phút cache
                return cached_res

        try:
            d_obj = datetime.strptime(t_date, "%Y-%m-%d").date()
        except Exception:
            d_obj = today_obj
            t_date = today_str

        day_str = f"{d_obj.day:02d}"
        month_str = f"{d_obj.month:02d}"
        display_date = d_obj.strftime("%d/%m/%Y")

        sheet_id = self.get_sheet_id_for_month(month_str)
        if not sheet_id:
            sheet_id = DEFAULT_SHEET_AUGUST_ID

        rows = []

        # Cách 1: Thử tải CSV nhanh qua GID nếu có sẵn
        if month_str == "08" and day_str in AUGUST_DAY_GIDS and sheet_id == DEFAULT_SHEET_AUGUST_ID:
            gid = AUGUST_DAY_GIDS[day_str]
            url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=10) as resp:
                    content = resp.read().decode("utf-8", errors="ignore")
                reader = csv.reader(io.StringIO(content))
                rows = list(reader)
            except Exception:
                rows = []

        # Cách 2: Đọc qua XLSX cached
        if not rows:
            try:
                data = self._get_xlsx_bytes(sheet_id)
                wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
                matching_sheet = None

                for s_name in wb.sheetnames:
                    clean_s = s_name.strip()
                    if clean_s.startswith(day_str) or clean_s.startswith(f"{d_obj.day} "):
                        matching_sheet = s_name
                        break

                if not matching_sheet:
                    for s_name in wb.sheetnames:
                        if day_str in s_name:
                            matching_sheet = s_name
                            break

                if matching_sheet and matching_sheet in wb.sheetnames:
                    ws = wb[matching_sheet]
                    for r in ws.iter_rows(values_only=True):
                        row_vals = [str(c).strip() if c is not None else "" for c in r]
                        if any(row_vals):
                            rows.append(row_vals)
            except Exception as e:
                print(f"Error fetching dynamic XLSX: {e}")

        # Phân tích dòng bảng kê
        ben_trips = []
        thung_trips = []
        current_section = "ben"

        for row in rows:
            if not row or len(row) < 2:
                continue
            row_joined = " ".join(row).upper()
            if "XE CÔNG DÀI" in row_joined or "XE THÙNG" in row_joined:
                current_section = "thung"
                continue
            elif "XE BEN" in row_joined and "TĂNG BO" not in row_joined:
                current_section = "ben"
                continue
            elif "TĂNG BO" in row_joined:
                break

            first_col = row[0].strip() if len(row) > 0 else ""
            raw_plate = row[1].strip() if len(row) > 1 else ""
            if not raw_plate or "BIỂN SỐ" in raw_plate.upper() or "SỐ XE" in raw_plate.upper() or "STT" in first_col.upper():
                continue

            # Chuẩn hóa biển số
            clean_p = raw_plate.replace(" ", "").replace("-", "").replace(".", "").upper()
            if not (clean_p.startswith("63") or clean_p.startswith("66")):
                continue

            if len(clean_p) == 8:
                formatted_plate = f"{clean_p[:3]}-{clean_p[3:6]}.{clean_p[6:]}"
            else:
                formatted_plate = raw_plate

            route = row[2].strip() if len(row) > 2 else ""
            cargo = row[3].strip() if len(row) > 3 else ""
            customer = row[4].strip() if len(row) > 4 else ""

            origin = ""
            dest = ""
            if "=>" in route:
                parts = route.split("=>")
                origin = parts[0].strip()
                dest = parts[1].strip()
            elif "->" in route:
                parts = route.split("->")
                origin = parts[0].strip()
                dest = parts[1].strip()
            else:
                dest = route

            combined_notes = (route + " " + customer + " " + cargo).upper()
            if "TÀI XẾ NGHỈ" in combined_notes or "TX NGHỈ" in combined_notes or "NGHỈ" in combined_notes:
                status_code = "driver_off"
                status_str = "Tài xế nghỉ cả ngày"
            elif "SỬA XE" in combined_notes or "BẢO DƯỠNG" in combined_notes or "SỬA" in combined_notes:
                status_code = "off"
                status_str = "Bảo dưỡng / Sửa xe"
            elif "TRỐNG LỊCH" in combined_notes or "ĐỢI GIAO" in combined_notes:
                status_code = "idle"
                status_str = "Đợi giao hàng / Trống lịch" if "ĐỢI GIAO" in combined_notes else "Trống lịch"
            elif not customer and not route:
                status_code = "idle"
                status_str = "Nghỉ bãi / Trống lịch"
            else:
                status_code = "active"
                status_str = "Đang hoạt động"

            driver_name = OFFICIAL_DRIVERS_BY_CLEAN_PLATE.get(clean_p, "Tài xế công ty")

            trip_item = {
                "raw_plate": formatted_plate,
                "plate_number": formatted_plate,
                "clean_plate": clean_p,
                "vehicle_type": "Xe Ben" if current_section == "ben" else "Xe Thùng",
                "route": route or ("Nghỉ bãi" if status_code != "active" else "Đang chạy"),
                "origin": origin,
                "destination": dest,
                "cargo_type": cargo or "—",
                "customer_name": customer or "—",
                "status_code": status_code,
                "status_text": status_str,
                "driver_name": driver_name,
                "num_trips": 1 if status_code == "active" else 0,
                "date": display_date
            }

            if current_section == "ben":
                ben_trips.append(trip_item)
            else:
                thung_trips.append(trip_item)

        all_trips = []
        for idx, t in enumerate(ben_trips + thung_trips, 1):
            t["stt"] = idx
            all_trips.append(t)

        total_vehicles = len(all_trips)
        active_count = len([t for t in all_trips if t.get("status_code") == "active"])
        off_count = len([t for t in all_trips if t.get("status_code") != "active"])
        active_percent = round((active_count / total_vehicles * 100), 1) if total_vehicles > 0 else 0.0

        # Phân tích cơ cấu chủ hàng (%)
        from collections import Counter
        cust_counter = Counter()
        for t in all_trips:
            if t.get("status_code") == "active":
                c = t.get("customer_name")
                if c and c != "—" and "NGHỈ" not in c.upper() and "SỬA" not in c.upper() and "TRỐNG" not in c.upper():
                    cust_counter[c] += 1

        total_cust_trips = sum(cust_counter.values())
        customer_breakdown = []
        for c_name, c_cnt in cust_counter.most_common(5):
            pct = round(c_cnt / total_cust_trips * 100, 1) if total_cust_trips > 0 else 0.0
            customer_breakdown.append({"customer": c_name, "trips": c_cnt, "percent": pct})

        result = {
            "date": display_date,
            "selected_date": t_date,
            "ben": ben_trips,
            "thung": thung_trips,
            "ben_trips": ben_trips,
            "thung_trips": thung_trips,
            "all_trips": all_trips,
            "trips": all_trips,
            "total_ben": len(ben_trips),
            "total_thung": len(thung_trips),
            "active_ben": len([t for t in ben_trips if t.get("status_code") == "active"]),
            "active_thung": len([t for t in thung_trips if t.get("status_code") == "active"]),
            "active_count": active_count,
            "off_count": off_count,
            "active_percent": active_percent,
            "customer_breakdown": customer_breakdown,
            "sync_time": datetime.now().strftime("%H:%M:%S")
        }

        # Lưu cache
        self.cached_trips[t_date] = (now, result)
        return result

    def get_weekly_dispatch_stats(self, target_date=None):
        """
        Tính toán chính xác hoạt động tuần (7 ngày) của toàn bộ 26 xe từ bảng kê Google Sheets trong 1 lượt đọc siêu tốc.
        Mỗi chuyến chạy có cự ly khứ hồi thực tế từ tuyến đường (trung bình 130km - 180km).
        """
        today_obj = date.today()
        if isinstance(target_date, str):
            try:
                base_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except Exception:
                base_date = today_obj
        elif isinstance(target_date, date):
            base_date = target_date
        else:
            base_date = today_obj

        month_str = f"{base_date.month:02d}"
        sheet_id = self.get_sheet_id_for_month(month_str) or DEFAULT_SHEET_AUGUST_ID

        days_to_fetch = [base_date - timedelta(days=i) for i in range(7)]
        day_prefixes = [f"{d.day:02d}" for d in days_to_fetch] + [str(d.day) for d in days_to_fetch]

        truck_stats = {}
        try:
            data = self._get_xlsx_bytes(sheet_id)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)

            sheet_to_date = {}
            for s in wb.sheetnames:
                cs = s.strip()
                for dp in day_prefixes:
                    if cs.startswith(dp + " ") or cs.startswith(dp + "-") or cs == dp:
                        sheet_to_date[s] = dp
                        break

            for s_name in sheet_to_date.keys():
                ws = wb[s_name]
                for row in ws.iter_rows(values_only=True):
                    if not row or len(row) < 2:
                        continue
                    first_col = str(row[0]).strip() if row[0] is not None else ""
                    if "XE BEN" in first_col.upper() or "XE CÔNG" in first_col.upper() or "STT" in first_col.upper():
                        continue

                    raw_plate = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                    if not raw_plate or "BIỂN" in raw_plate.upper() or "SỐ XE" in raw_plate.upper():
                        continue

                    clean_p = raw_plate.replace("-", "").replace(".", "").replace(" ", "").upper()
                    if len(clean_p) < 7:
                        continue

                    trailer = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                    driver = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                    customer = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                    origin = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
                    dest = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                    cargo = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                    num_trips_raw = str(row[8]).strip() if len(row) > 8 and row[8] is not None else "1"

                    try:
                        num_trips = int(float(num_trips_raw))
                    except Exception:
                        num_trips = 1

                    if clean_p not in truck_stats:
                        truck_stats[clean_p] = {
                            "plate_number": raw_plate,
                            "trailer_number": trailer,
                            "driver_name": driver,
                            "active_days": 0,
                            "total_trips": 0,
                            "routes": set(),
                            "estimated_km": 0.0
                        }

                    lower_dest = dest.lower() + " " + customer.lower() + " " + cargo.lower()
                    is_active = not ("nghỉ" in lower_dest or "bảo dưỡng" in lower_dest or "sửa" in lower_dest or (not customer and not dest))

                    if is_active:
                        leg_km = 140.0
                        if "đức hòa" in dest.lower() or "bến lức" in dest.lower() or "long an" in dest.lower():
                            leg_km = 180.0
                        elif "thăng long" in dest.lower() or "đan phượng" in dest.lower() or "bình dương" in dest.lower():
                            leg_km = 150.0
                        elif "cát lái" in dest.lower() or "hiệp phước" in dest.lower() or "hcm" in dest.lower():
                            leg_km = 130.0
                        elif "cảng" in dest.lower() or "ptsc" in dest.lower() or "mỹ xuân" in dest.lower():
                            leg_km = 80.0
                        elif "vũng tàu" in dest.lower():
                            leg_km = 90.0

                        trip_km = leg_km * num_trips
                        truck_stats[clean_p]["total_trips"] += num_trips
                        truck_stats[clean_p]["active_days"] += 1
                        truck_stats[clean_p]["estimated_km"] += trip_km
                        if origin and dest:
                            truck_stats[clean_p]["routes"].add(f"{origin} ➔ {dest}")

        except Exception as e:
            print(f"Error calculating weekly stats: {e}")

        return truck_stats

    def fetch_maintenance_sheet(self):
        now = time.time()
        if self.cached_maintenance:
            ts, data = self.cached_maintenance
            if now - ts < 120:
                return data

        try:
            req = urllib.request.Request(SHEET2_MAINTENANCE_URL, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=12) as resp:
                content = resp.read().decode("utf-8", errors="ignore")

            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            maintenance_list = []
            count = 0
            for r in rows:
                if len(r) > 2 and r[1].strip().isdigit():
                    stt = int(r[1].strip())
                    if stt != count + 1 or count >= 25:
                        continue
                    count += 1

                    raw_plate = r[2].strip()
                    clean_p = raw_plate.replace(" ", "").replace("-", "").replace(".", "").upper()
                    if len(clean_p) == 8:
                        plate_formatted = f"{clean_p[:3]}-{clean_p[3:6]}.{clean_p[6:]}"
                    else:
                        plate_formatted = raw_plate

                    norm = r[3].strip() if len(r) > 3 and r[3].strip() else "15.000"
                    last_km = r[4].strip() if len(r) > 4 and r[4].strip() else "—"
                    last_date = r[5].strip() if len(r) > 5 and r[5].strip() else "—"
                    current_km = r[6].strip() if len(r) > 6 and r[6].strip() else "—"
                    due_km = r[8].strip() if len(r) > 8 and r[8].strip() else "—"
                    remaining_km = r[9].strip() if len(r) > 9 and r[9].strip() else "—"
                    status_text = r[10].strip() if len(r) > 10 and r[10].strip() else ""
                    notes = r[11].strip() if len(r) > 11 else ""

                    if not status_text:
                        if remaining_km and remaining_km != "—":
                            try:
                                rem_num = float(remaining_km.replace(".", "").replace(",", "."))
                                if rem_num < 0:
                                    status_text = "Sắp / quá hạn"
                                elif rem_num < 3000:
                                    status_text = "Gần tới hạn"
                                else:
                                    status_text = "Còn xa"
                            except Exception:
                                status_text = "Còn xa"
                        else:
                            status_text = "Thiếu số liệu"

                    maintenance_list.append({
                        "stt": stt,
                        "plate_number": plate_formatted,
                        "raw_plate": raw_plate,
                        "norm_km": norm,
                        "last_km": last_km,
                        "last_date": last_date,
                        "current_km": current_km,
                        "due_km": due_km,
                        "remaining_km": remaining_km,
                        "status": status_text,
                        "status_text": status_text,
                        "notes": notes
                    })

            if maintenance_list:
                self.cached_maintenance = (now, maintenance_list)
                return maintenance_list
        except Exception as e:
            print(f"Error fetching maintenance sheet: {e}")
            if self.cached_maintenance:
                return self.cached_maintenance[1]
        return []

    def fetch_vehicles_sheet(self):
        now = time.time()
        if self.cached_vehicles:
            ts, data = self.cached_vehicles
            if now - ts < 300:
                return data

        try:
            req = urllib.request.Request(SHEET2_VEHICLES_URL, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=12) as resp:
                content = resp.read().decode("utf-8", errors="ignore")

            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            ben_vehicles = []
            thung_vehicles = []
            is_thung = False

            for r in rows:
                if not r or len(r) < 2:
                    continue
                first = r[0].strip().upper()
                if "XE CÔNG DÀI" in first or "XE THÙNG" in first:
                    is_thung = True
                    continue
                if "XE BEN" in first:
                    is_thung = False
                    continue

                if len(r) > 2 and r[1].strip().isdigit():
                    v_item = {
                        "stt": int(r[1].strip()),
                        "plate_number": r[2].strip() if len(r) > 2 else "",
                        "trailer_number": r[3].strip() if len(r) > 3 else "",
                        "driver_name": r[4].strip() if len(r) > 4 else "",
                        "phone": r[5].strip() if len(r) > 5 else "",
                        "driver_status": r[6].strip() if len(r) > 6 else "Bình thường"
                    }
                    if is_thung:
                        thung_vehicles.append(v_item)
                    else:
                        ben_vehicles.append(v_item)

            res = {"ben": ben_vehicles, "thung": thung_vehicles}
            self.cached_vehicles = (now, res)
            return res
        except Exception as e:
            print(f"Error fetching vehicles sheet: {e}")
            if self.cached_vehicles:
                return self.cached_vehicles[1]
            return {"ben": [], "thung": []}

sheets_client = GoogleSheetsSyncService()
