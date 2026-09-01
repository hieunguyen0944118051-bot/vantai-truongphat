from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import models, schemas, auth
from database import get_db
from sheets_service import sheets_client
from gps_service import gps_client

router = APIRouter(prefix="/api/trips", tags=["Trips"])

@router.get("/sheets-live")
def get_sheets_live_trips(
    view_date: Optional[str] = Query(None, alias="date"),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Get live daily dispatch from Google Sheets, merged with live BA GPS telemetry.
    """
    try:
        res = sheets_client.fetch_daily_trips(target_date=view_date)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi đọc Google Sheets: {str(e)}")

    # Merge real-time GPS telemetry directly
    try:
        fleet = gps_client.fetch_live_fleet()
        gps_map = {
            v["plate_number"].replace("-", "").replace(".", "").replace(" ", "").upper(): v
            for v in fleet
        }
    except Exception:
        gps_map = {}

    for t in res["all_trips"]:
        clean_p = t["raw_plate"].replace("-", "").replace(".", "").replace(" ", "").upper()
        tel = gps_map.get(clean_p)
        if tel:
            t["gps_address"] = tel["address"] or "Bãi xe Trường Phát"
            t["speed"] = tel["speed"]
            t["is_running"] = tel["speed"] > 0
            if tel["speed"] > 0:
                t["movement_state"] = f"🟢 Xe đang chạy ({tel['speed']} km/h)"
                t["op_badge_class"] = "bg-emerald-100 text-emerald-800"
            elif tel["status_type"] == "idling":
                t["movement_state"] = "🟡 Xe dừng nổ máy"
                t["op_badge_class"] = "bg-amber-100 text-amber-800"
            else:
                t["movement_state"] = "⚪ Xe đang đỗ (Tắt máy)"
                t["op_badge_class"] = "bg-slate-100 text-slate-700"
        else:
            t["gps_address"] = "Bãi xe Trường Phát / Đậu bãi"
            t["speed"] = 0
            t["is_running"] = False
            t["movement_state"] = "⚪ Xe đậu bãi"
            t["op_badge_class"] = "bg-slate-100 text-slate-600"

    return {"success": True, "data": res}

@router.post("/sync-sheets")
def sync_trips_from_sheets(
    view_date: Optional[str] = Query(None, alias="date"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role(["admin", "manager"]))
):
    """
    Sync trips from Google Sheets into the local database for historical reporting.
    """
    try:
        if view_date:
            try:
                sync_date = datetime.strptime(view_date, "%Y-%m-%d").date()
            except Exception:
                sync_date = date.today()
        else:
            sync_date = date.today()

        data = sheets_client.fetch_daily_trips(target_date=sync_date.strftime("%Y-%m-%d"))
        trips = data.get("all_trips", [])
        
        vehicles = db.query(models.Vehicle).all()
        v_map = {v.plate_number.replace("-", "").replace(".", "").replace(" ", "").upper(): v for v in vehicles}

        date_prefix = sync_date.strftime("%Y%m%d")
        synced_count = 0

        for t in trips:
            clean_plate = t["raw_plate"].replace("-", "").replace(".", "").replace(" ", "").upper()
            v = v_map.get(clean_plate)
            
            trip_code = f"TRIP-{date_prefix}-{t['stt']:03d}-{clean_plate[-4:]}"
            existing = db.query(models.Trip).filter(
                models.Trip.trip_code == trip_code
            ).first()

            if existing:
                existing.customer_name = t["customer_name"]
                existing.cargo_type = t["cargo_type"]
                existing.origin = t["origin"]
                existing.destination = t["destination"]
                existing.num_trips = 1 if t["status_code"] == "active" else 0
                existing.weight_tons = 30.0 if t["status_code"] == "active" else 0.0
                existing.notes = t["route"]
            else:
                new_trip = models.Trip(
                    trip_date=sync_date,
                    trip_code=trip_code,
                    vehicle_id=v.id if v else None,
                    customer_name=t["customer_name"],
                    cargo_type=t["cargo_type"],
                    origin=t["origin"],
                    destination=t["destination"],
                    num_trips=1 if t["status_code"] == "active" else 0,
                    weight_tons=30.0 if t["status_code"] == "active" else 0.0,
                    notes=t["route"]
                )
                db.add(new_trip)
                synced_count += 1
        db.commit()
        return {
            "success": True,
            "message": f"Đã đồng bộ thành công {len(trips)} chuyến từ Google Trang Tính ngày {sync_date.strftime('%d/%m/%Y')}!",
            "active_count": data.get("active_count", 0),
            "off_count": data.get("off_count", 0),
            "synced_count": synced_count
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi khi đồng bộ Google Sheets: {str(e)}")

@router.get("", response_model=List[schemas.TripOut])
def get_trips(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    vehicle_id: Optional[int] = None,
    customer_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    query = db.query(models.Trip)
    if date_from:
        query = query.filter(models.Trip.trip_date >= date_from)
    if date_to:
        query = query.filter(models.Trip.trip_date <= date_to)
    if vehicle_id:
        query = query.filter(models.Trip.vehicle_id == vehicle_id)
    if customer_name:
        query = query.filter(models.Trip.customer_name.ilike(f"%{customer_name}%"))
    return query.order_by(models.Trip.trip_date.desc(), models.Trip.id.desc()).all()

@router.get("/export/excel")
def export_trips_excel(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    try:
        data = sheets_client.fetch_daily_trips()
        trips_data = data["all_trips"]
    except Exception:
        trips_data = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lệnh Điều Xe Thực Tế"

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "CÔNG TY TNHH DV VẬN TẢI TRƯỜNG PHÁT - BẢNG ĐIỀU PHỐI XE THỰC TẾ"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "STT", "Biển Số Xe", "Loại Xe", "Tuyến Đường (Đi => Đến)",
        "Loại Hàng Hóa", "Chủ Hàng", "Trạng Thái Hoạt Động"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE')
    )

    for row_idx, t in enumerate(trips_data, 3):
        ws.cell(row=row_idx, column=1, value=t["stt"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=t["plate_number"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=t["vehicle_type"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=t["route"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=5, value=t["cargo_type"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=6, value=t["customer_name"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=7, value=t["status_text"]).alignment = Alignment(horizontal="center")

        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).border = thin_border
            ws.cell(row=row_idx, column=c).font = Font(name="Arial", size=10)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Bang_Ke_Dieu_Xe_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/sheets-config")
def get_sheets_config(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    setting_08 = db.query(models.SystemSetting).filter(models.SystemSetting.key == "sheet_url_month_08").first()
    setting_09 = db.query(models.SystemSetting).filter(models.SystemSetting.key == "sheet_url_month_09").first()
    return {
        "sheet_url_month_08": setting_08.value if setting_08 else "https://docs.google.com/spreadsheets/d/1p0B1bx_yUM6BfW2D88P-Jgra3sBSfqHEM_Op35WxSpI/edit",
        "sheet_url_month_09": setting_09.value if setting_09 else f"https://docs.google.com/spreadsheets/d/{sheets_service.DEFAULT_SHEET_SEPTEMBER_ID}/edit"
    }

@router.post("/sheets-config")
def save_sheets_config(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role(["admin", "manager"]))
):
    month = str(payload.get("month", "09")).zfill(2)
    url = payload.get("url", "").strip()
    if url:
        sheets_client.set_sheet_url_for_month(month, url)
    return {"success": True, "message": f"Đã lưu thành công link Trang tính Tháng {month}!"}
